"""
analisis_estanteria_hybrid.py — Pipeline Híbrido SAM 3 + VLM

Arquitectura:
    Imagen → SAM 3 (text prompt: "product on supermarket shelf")
                → [mask_1, mask_2, ..., mask_n]  (un mask por producto detectado)
                                    ↓
                  VLM individual: "qué es este producto + precio"
                                    ↓
                       [{name, price, qty:1}, ...]

Dependencias:
    pip install -r requirements_sam.txt
    git clone https://github.com/facebookresearch/sam3.git && cd sam3 && pip install -e .
    # Modelo gated: requiere HF_TOKEN (https://huggingface.co/facebook/sam3)

Ventajas vs pipeline original v4:
    - SAM 3 con text prompt detecta solo productos (no estantes, separadores, fondo)
    - El VLM NO cuenta, solo identifica → elimina sobre-conteo
    - Cada mask = 1 producto → conteo exacto por SAM
    - JSON pequeño por producto → sin truncamiento
    - Paralelización trivial: enviar masks al VLM en paralelo

Ventajas vs SAM 2.1:
    - Text prompts en lugar de segmentación ciega (AutomaticMaskGenerator)
    - Detección open-vocabulary por instancia (un mask+ID por producto)
    - Menos ruido: no genera masks de estantes, separadores, fondo
    - Mejor deduplicación: menos masks superpuestos del mismo producto
"""

import io
import json
import os
import random
import time
import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import numpy as np
from PIL import Image
import torch

from openai import OpenAI

# =============================================================================
# SAM 3 — Configuración
# =============================================================================
SAM3_BPE_PATH = None  # Se resuelve automáticamente desde el repo clonado
SAM3_CONFIDENCE = 0.1
SAM3_TEXT_PROMPT = "product"


@dataclass
class ProductResult:
    name: str
    price: Optional[float]
    quantity: int = 1
    confidence: float = 0.0
    mask_id: int = 0
    bbox: tuple[int, int, int, int] = (0, 0, 0, 0)


@dataclass
class ShelfAnalysisResult:
    image_name: str
    products: list[ProductResult] = field(default_factory=list)
    total_products: int = 0
    total_unique_products: int = 0
    sam_masks_generated: int = 0
    vlm_calls: int = 0
    errors: list[str] = field(default_factory=list)
    processing_time: float = 0.0


# =============================================================================
# SAM 3 — Funciones de segmentación
# =============================================================================


def get_sam_device() -> str:
    """Detecta el device disponible: cuda > cpu. SAM 3 no soporta MPS."""
    if torch.cuda.is_available():
        return "cuda"
    return "cpu"


def load_sam3_model(
    device: str = None,
    bpe_path: str = None,
    confidence_threshold: float = SAM3_CONFIDENCE,
):
    """
    Carga el modelo SAM 3 y retorna el processor.

    Args:
        device: "cuda" o "cpu". Si None, se autodetecta.
        bpe_path: Ruta al archivo BPE vocab. Si None, usa el default del repo.
        confidence_threshold: Umbral de confianza para filtrar detecciones.

    Returns:
        Sam3Processor configurado y listo para usar.
    """
    from sam3 import build_sam3_image_model
    from sam3.model.sam3_image_processor import Sam3Processor

    if device is None:
        device = get_sam_device()

    if device == "cuda" and not torch.cuda.is_available():
        print("[WARN] CUDA no disponible, usando CPU")
        device = "cpu"

    print(f"[INFO] SAM 3 usando device: {device}")

    build_kwargs = {"device": device}
    if bpe_path:
        build_kwargs["bpe_path"] = bpe_path

    model = build_sam3_image_model(**build_kwargs)
    processor = Sam3Processor(model, confidence_threshold=confidence_threshold)

    return processor


def generate_sam3_masks(
    processor,
    image: Image.Image,
    text_prompt: str = SAM3_TEXT_PROMPT,
    min_area: int = 500,
    max_masks: int = 100,
) -> list[tuple[np.ndarray, tuple[int, int, int, int], float]]:
    """
    Genera máscaras de productos usando SAM 3 con text prompt.

    A diferencia de SAM 2.1 (segmentación ciega), SAM 3 acepta un prompt
    de texto y retorna masks solo de objetos que coinciden con el concepto.

    Args:
        processor: Sam3Processor cargado
        image: PIL Image en RGB
        text_prompt: Concepto a buscar (ej: "product on supermarket shelf")
        min_area: Área mínima del mask en píxeles (filtra ruido)
        max_masks: Máximo número de masks a retornar

    Returns:
        Lista de (mask, bbox, score) donde:
            mask: np.ndarray booleano (H, W)
            bbox: (x1, y1, x2, y2) en coordenadas de imagen
            score: confianza de la detección (0.0-1.0)
    """
    # SAM 3 usa bfloat16 internamente. En CPU necesitamos autocast
    # para evitar mismatch de dtypes entre pesos (bf16) e inputs (f32).
    autocast_ctx = torch.autocast("cpu", dtype=torch.bfloat16) if not torch.cuda.is_available() else torch.autocast("cuda", dtype=torch.bfloat16)
    with autocast_ctx:
        state = processor.set_image(image)
        state = processor.set_text_prompt(text_prompt, state)

    masks = state.get("masks", [])
    boxes = state.get("boxes", [])
    scores = state.get("scores", [])

    if masks is None or len(masks) == 0:
        return []

    filtered = []

    for i, (mask_tensor, box, score) in enumerate(zip(masks, boxes, scores)):
        mask_np = mask_tensor[0].cpu().numpy().astype(bool)

        area = mask_np.sum()
        if area < min_area:
            continue

        score_val = float(score)

        # boxes de SAM 3 vienen en formato [x0, y0, x1, y1]
        x1, y1, x2, y2 = [int(v) for v in box[:4]]
        bbox = (x1, y1, x2, y2)

        # Log del primer box para verificar formato
        if i == 0 and len(filtered) == 0:
            print(f"  [DEBUG] Primer box SAM 3: {bbox}, score: {score_val:.3f}, area: {area}px")

        filtered.append((mask_np, bbox, score_val))

        if len(filtered) >= max_masks:
            break

    return filtered


def crop_mask_region(
    image: Image.Image,
    mask: np.ndarray,
    bbox: tuple[int, int, int, int],
    margin: int = 10,
) -> Image.Image:
    """
    Recorta la región de la imagen correspondiente a un mask.

    Args:
        image: Imagen original PIL
        mask: Máscara booleana numpy (H, W)
        bbox: (x1, y1, x2, y2) — formato directo de SAM 3
        margin: Pixeles extra de margen alrededor del crop

    Returns:
        PIL Image recortada con el producto, o None si el crop falla
    """
    x1, y1, x2, y2 = bbox
    h, w = mask.shape

    x1_c = max(0, x1 - margin)
    y1_c = max(0, y1 - margin)
    x2_c = min(w, x2 + margin)
    y2_c = min(h, y2 + margin)

    mask_crop = mask[y1_c:y2_c, x1_c:x2_c]

    if not mask_crop.any():
        return None

    pil_crop = image.crop((x1_c, y1_c, x2_c, y2_c))

    mask_img = Image.fromarray((mask_crop * 255).astype(np.uint8), mode="L")
    mask_resized = mask_img.resize(pil_crop.size, Image.NEAREST)

    result = Image.new("RGBA", pil_crop.size, (0, 0, 0, 0))
    result.paste(pil_crop, (0, 0))
    result.putalpha(mask_resized)

    result_rgb = result.convert("RGB")

    alpha_arr = np.array(mask_resized)
    if alpha_arr.max() == 0:
        return None

    mask_bool_u = alpha_arr > 127
    rows = np.any(mask_bool_u, axis=1)
    cols = np.any(mask_bool_u, axis=0)
    if not rows.any() or not cols.any():
        return None

    rmin, rmax = np.where(rows)[0][[0, -1]]
    cmin, cmax = np.where(cols)[0][[0, -1]]

    cropped = result_rgb.crop((cmin, rmin, cmax + 1, rmax + 1))

    return cropped


# =============================================================================
# VLM — Llamadas individuales por producto
# =============================================================================

SYSTEM_PROMPT_VLM = """Eres un asistente que identifica productos en imágenes de estanterías de supermercado.

Devuelve SOLO un JSON con esta estructura exacta, sin texto adicional:
{"name": "nombre del producto", "price": 12.99, "confidence": 0.95}

Reglas:
- name: nombre corto del producto (máx 5 palabras)
- price: precio del producto en la etiqueta visible, o null si no se ve
- confidence: número entre 0.0 y 1.0 indicando qué tan seguro estás
- Si no hay producto válido en la imagen, devuelve: {"name": "desconocido", "price": null, "confidence": 0.1}
- NO cuentes productos, solo identifica ESTE producto individual
- Sé conciso, el nombre debe tener máximo 5 palabras"""


def encode_image_to_base64(image: Image.Image) -> str:
    """Convierte PIL Image a base64 JPEG."""
    buffer = io.BytesIO()
    image.save(buffer, format="JPEG", quality=85)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def classify_single_product(
    image: Image.Image,
    client: OpenAI,
    model: str = "qwen2.5vl:7b",
    temperature: float = 0.1,
) -> ProductResult:
    """
    Clasifica un solo producto (recortado por SAM) usando el VLM.

    Args:
        image: PIL Image del producto recortado
        client: OpenAI client conectado a Ollama
        model: Modelo VLM a usar
        temperature: Temperatura para generación

    Returns:
        ProductResult con name, price, qty=1, confidence
    """
    img_b64 = encode_image_to_base64(image)

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT_VLM},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"},
                        },
                        {
                            "type": "text",
                            "text": "Identifica este producto de la estantería.",
                        },
                    ],
                },
            ],
            temperature=temperature,
            max_tokens=150,
        )

        raw = response.choices[0].message.content.strip()

        if raw.startswith("```json"):
            raw = raw[7:]
        if raw.startswith("```"):
            raw = raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3]

        data = json.loads(raw.strip())

        return ProductResult(
            name=data.get("name", "desconocido"),
            price=data.get("price"),
            quantity=1,
            confidence=data.get("confidence", 0.5),
        )

    except Exception as e:
        print(f"  [WARN] Error clasificando producto: {type(e).__name__}: {e}")
        return ProductResult(
            name="error_parse",
            price=None,
            quantity=1,
            confidence=0.0,
        )


# =============================================================================
# PIPELINE PRINCIPAL
# =============================================================================


def run_hybrid_pipeline(
    image_path: str,
    sam_processor,
    vlm_client: OpenAI,
    vlm_model: str = "qwen2.5vl:7b",
    text_prompt: str = SAM3_TEXT_PROMPT,
    max_workers: int = 10,
    min_mask_area: int = 500,
) -> ShelfAnalysisResult:
    """
    Pipeline híbrido SAM 3 + VLM para análisis de estanterías.

    Flujo:
        1. SAM 3 genera máscaras con text prompt (solo productos)
        2. Cada máscara se recorta → 1 crop por producto
        3. Cada crop se envía al VLM en paralelo (ThreadPoolExecutor)
        4. Agregar resultados

    Args:
        image_path: Ruta a la imagen JPG
        sam_processor: Sam3Processor cargado
        vlm_client: OpenAI client conectado a Ollama
        vlm_model: Modelo VLM a usar
        text_prompt: Concepto para SAM 3
        max_workers: Threads paralelos para llamadas VLM
        min_mask_area: Área mínima de mask en píxeles

    Returns:
        ShelfAnalysisResult con lista de productos identificados
    """
    start_time = time.time()
    image_name = Path(image_path).name

    result = ShelfAnalysisResult(image_name=image_name)
    errors = []

    try:
        image = Image.open(image_path).convert("RGB")
    except Exception as e:
        errors.append(f"Error al abrir imagen: {e}")
        result.errors = errors
        return result

    masks_with_data = generate_sam3_masks(
        sam_processor, image, text_prompt=text_prompt, min_area=min_mask_area, max_masks=100
    )
    result.sam_masks_generated = len(masks_with_data)

    if not masks_with_data:
        errors.append("SAM 3 no detectó ningún producto")
        result.errors = errors
        result.processing_time = time.time() - start_time
        return result

    def process_single_mask(args):
        idx, (mask, bbox, score) = args
        crop = crop_mask_region(image, mask, bbox, margin=10)
        if crop is None:
            return None
        vlm_result = classify_single_product(crop, vlm_client, model=vlm_model)
        vlm_result.mask_id = idx
        vlm_result.bbox = bbox
        vlm_result.confidence = max(vlm_result.confidence, score)
        return vlm_result

    vlm_results: list[ProductResult] = []
    vlm_calls = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(process_single_mask, (i, md)): i
            for i, md in enumerate(masks_with_data)
        }

        for future in as_completed(futures):
            vlm_calls += 1
            try:
                res = future.result()
                if res is not None:
                    vlm_results.append(res)
            except Exception as e:
                errors.append(f"Error en mask {futures[future]}: {e}")

    result.products = vlm_results
    result.total_products = len(vlm_results)
    result.vlm_calls = vlm_calls
    result.errors = errors
    result.processing_time = time.time() - start_time

    return result


# =============================================================================
# DEDUPLICACIÓN post-VLM
# =============================================================================


def deduplicate_products(
    products: list[ProductResult], similarity_thresh: float = 0.85
) -> list[ProductResult]:
    """
    Deduplica productos identificados múltiples veces por SAM.

    Aunque SAM 3 con text prompts genera menos duplicados que SAM 2.1,
    masks superpuestos del mismo producto siguen siendo posibles. Usa:
    - Similitud de nombre (lowercase, sin números)
    - Coincidencia de precio
    """
    if not products:
        return []

    deduped = []
    used = set()

    products_sorted = sorted(products, key=lambda p: p.confidence, reverse=True)

    for p in products_sorted:
        if p.name == "desconocido" or p.name == "error_parse":
            continue

        is_duplicate = False
        p_name_norm = p.name.lower().strip()

        for used_idx in used:
            other = deduped[used_idx]
            other_norm = other.name.lower().strip()

            name_similar = p_name_norm in other_norm or other_norm in p_name_norm
            price_match = (
                p.price is not None
                and other.price is not None
                and abs(p.price - other.price) < 0.01
            )

            if name_similar and price_match:
                is_duplicate = True
                break

        if not is_duplicate:
            deduped.append(p)
            used.add(len(deduped) - 1)

    return deduped


# =============================================================================
# EXPORTACIÓN A EXCEL
# =============================================================================


def export_to_excel(results: list[ShelfAnalysisResult], output_path: str) -> None:
    """Exporta resultados a Excel con 3 hojas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_detalle = wb.create_sheet("Detalle de Productos")
    ws_stats = wb.create_sheet("Estadísticas")

    header_fill = PatternFill(
        start_color="3D008D", end_color="ED1E79", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Hoja Resumen ---
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 15
    ws_summary.column_dimensions["E"].width = 15
    ws_summary.column_dimensions["F"].width = 12

    headers_s = [
        "Imagen",
        "Total Productos",
        "Masks SAM",
        "Llamadas VLM",
        "Tiempo (s)",
        "Errores",
    ]
    for col, h in enumerate(headers_s, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for row, r in enumerate(results, 2):
        ws_summary.cell(row=row, column=1, value=r.image_name).border = thin_border
        ws_summary.cell(row=row, column=2, value=r.total_products).border = thin_border
        ws_summary.cell(
            row=row, column=3, value=r.sam_masks_generated
        ).border = thin_border
        ws_summary.cell(row=row, column=4, value=r.vlm_calls).border = thin_border
        ws_summary.cell(
            row=row, column=5, value=round(r.processing_time, 1)
        ).border = thin_border
        ws_summary.cell(
            row=row, column=6, value=", ".join(r.errors) if r.errors else "-"
        ).border = thin_border

    # --- Hoja Detalle ---
    headers_d = ["Imagen", "Producto", "Precio", "Cantidad", "Confianza", "Mask ID"]
    for col, h in enumerate(headers_d, 1):
        cell = ws_detalle.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for r in results:
        for p in r.products:
            ws_detalle.cell(row=row, column=1, value=r.image_name).border = thin_border
            ws_detalle.cell(row=row, column=2, value=p.name).border = thin_border
            ws_detalle.cell(row=row, column=3, value=p.price).border = thin_border
            ws_detalle.cell(row=row, column=4, value=p.quantity).border = thin_border
            ws_detalle.cell(
                row=row, column=5, value=round(p.confidence, 2)
            ).border = thin_border
            ws_detalle.cell(row=row, column=6, value=p.mask_id).border = thin_border
            row += 1

    for col in range(1, 7):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 18

    # --- Hoja Estadísticas ---
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 15

    headers_st = ["Métrica", "Valor"]
    for col, h in enumerate(headers_st, 1):
        cell = ws_stats.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    total_imgs = len(results)
    total_products = sum(r.total_products for r in results)
    total_masks = sum(r.sam_masks_generated for r in results)
    total_vlm_calls = sum(r.vlm_calls for r in results)
    avg_time = sum(r.processing_time for r in results) / total_imgs if total_imgs else 0
    avg_products_per_image = total_products / total_imgs if total_imgs else 0

    stats = [
        ("Total imágenes procesadas", total_imgs),
        ("Total productos detectados", total_products),
        ("Total masks SAM generados", total_masks),
        ("Total llamadas VLM", total_vlm_calls),
        ("Tiempo promedio por imagen (s)", round(avg_time, 1)),
        ("Productos promedio por imagen", round(avg_products_per_image, 1)),
        (
            "SAM masks / producto VLM",
            round(total_masks / total_vlm_calls, 2) if total_vlm_calls else 0,
        ),
    ]

    for row_idx, (metric, value) in enumerate(stats, 2):
        ws_stats.cell(row=row_idx, column=1, value=metric).border = thin_border
        ws_stats.cell(row=row_idx, column=2, value=value).border = thin_border

    wb.save(output_path)


# =============================================================================
# MAIN — Ejecución CLI
# =============================================================================


def main():
    import argparse
    import glob

    parser = argparse.ArgumentParser(description="Pipeline híbrido SAM 3 + VLM")
    parser.add_argument("--image", type=str, help="Ruta a imagen individual")
    parser.add_argument(
        "--sample",
        type=int,
        default=10,
        help="Porcentaje de imágenes a muestrear del archive (0-100)",
    )
    parser.add_argument(
        "--archive", type=str, default="/archive/images", help="Carpeta con imágenes"
    )
    parser.add_argument(
        "--output",
        type=str,
        default="/data/resultados/analisis_hybrid",
        help="Carpeta de salida para Excel/JSON",
    )
    parser.add_argument(
        "--sam-device",
        type=str,
        default=None,
        choices=["cuda", "cpu"],
        help="Device para SAM 3 (default: autodetectar). SAM 3 no soporta MPS.",
    )
    parser.add_argument(
        "--max-workers", type=int, default=10, help="Threads paralelos para VLM"
    )
    parser.add_argument(
        "--model", type=str, default="qwen2.5vl:7b", help="Modelo VLM en Ollama"
    )
    parser.add_argument(
        "--ollama-url",
        type=str,
        default=os.environ.get("OLLAMA_HOST", "http://host.docker.internal:11434/v1"),
        help="URL base de Ollama (OpenAI-compatible, necesita /v1). "
             "Se lee de OLLAMA_HOST env var si existe.",
    )
    parser.add_argument(
        "--text-prompt",
        type=str,
        default=SAM3_TEXT_PROMPT,
        help="Text prompt para SAM 3 (concepto a detectar)",
    )
    parser.add_argument(
        "--confidence",
        type=float,
        default=SAM3_CONFIDENCE,
        help="Umbral de confianza para SAM 3 (0.0-1.0)",
    )
    parser.add_argument(
        "--bpe-path",
        type=str,
        default=None,
        help="Ruta al archivo BPE vocab de SAM 3 (default: auto del repo)",
    )

    args = parser.parse_args()

    Path(args.output).mkdir(parents=True, exist_ok=True)

    # Device: respetar argumento del usuario, o autodetectar
    sam_device = args.sam_device if args.sam_device else get_sam_device()
    print(f"[INFO] Cargando SAM 3 en device: {sam_device}")

    sam_processor = load_sam3_model(
        device=sam_device,
        bpe_path=args.bpe_path,
        confidence_threshold=args.confidence,
    )

    vlm_client = OpenAI(base_url=args.ollama_url, api_key="ollama")

    if args.image:
        image_paths = [args.image]
    else:
        all_images = sorted(glob.glob(f"{args.archive}/*.jpg"))
        sample_size = max(1, int(len(all_images) * args.sample / 100))
        random.seed(42)
        image_paths = random.sample(all_images, sample_size)
        print(
            f"[INFO] Procesando {len(image_paths)} imágenes ({args.sample}% del archive)"
        )

    results: list[ShelfAnalysisResult] = []

    for img_path in image_paths:
        print(f"[INFO] Procesando: {img_path}")
        result = run_hybrid_pipeline(
            img_path,
            sam_processor,
            vlm_client,
            vlm_model=args.model,
            text_prompt=args.text_prompt,
            max_workers=args.max_workers,
            min_mask_area=500,
        )

        result.products = deduplicate_products(result.products)
        result.total_unique_products = len(result.products)

        results.append(result)

        print(
            f"  -> {result.total_products} productos, "
            f"{result.sam_masks_generated} masks SAM, "
            f"{result.vlm_calls} llamadas VLM, "
            f"{result.processing_time:.1f}s"
        )

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    excel_path = f"{args.output}/analisis_hybrid_{timestamp}.xlsx"
    json_path = f"{args.output}/analisis_hybrid_{timestamp}.json"

    export_to_excel(results, excel_path)
    print(f"[INFO] Excel guardado: {excel_path}")

    with open(json_path, "w") as f:
        json.dump(
            [
                {
                    "image_name": r.image_name,
                    "products": [asdict(p) for p in r.products],
                    "total_products": r.total_products,
                    "sam_masks_generated": r.sam_masks_generated,
                    "vlm_calls": r.vlm_calls,
                    "errors": r.errors,
                    "processing_time": round(r.processing_time, 2),
                }
                for r in results
            ],
            f,
            indent=2,
        )
    print(f"[INFO] JSON guardado: {json_path}")


if __name__ == "__main__":
    main()
