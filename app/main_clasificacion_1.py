import os
import fitz  # PyMuPDF
import io
import re
import json
import cv2
import numpy as np
import glob
import base64
from PIL import Image, ImageEnhance
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from deskew import determine_skew
from openai import OpenAI
from typing import Dict, Any, Union

# --- CONFIGURACIÓN GLOBAL ---
MODEL_NAME = "llama3.2-vision" 
DPI = 150  # 150 DPI es el sweet spot para OCR/Visión sin generar payloads gigantes
OLLAMA_URL = os.getenv('OLLAMA_HOST', 'http://host.docker.internal:11434/v1')

client = OpenAI(base_url=OLLAMA_URL, api_key='ollama')

# --- DEFINICIONES DE NEGOCIO ---
# Se definen aquí para fácil mantenimiento. Serán inyectadas dinámicamente en el prompt.
DEFINICIONES_CATEGORIAS = """
1. QUEJA_RECLAMO: Documentos donde un ciudadano expresa inconformidad. Palabras clave: "Petición", "Reclamo", "Solicito", "Inconformidad", "Derecho de petición".
2. CONTRATO: Acuerdos legales entre partes. Palabras clave: "Contratante", "Contratista", "Cláusulas", "Objeto del contrato", "Prestación de servicios", "Firma".
3. RESOLUCION_ADMINISTRATIVA: Actos administrativos formales. Palabras clave: "RESOLUCION No", "RESUELVE", "CONSIDERANDO", "ARTICULO PRIMERO", "COMUNIQUESE Y CUMPLASE".
4. INFORME_TECNICO: Reportes de avance, diagnósticos o resultados. Palabras clave: "Informe de gestión", "Resultados", "Diagnóstico", "Objetivo", "Alcance", "Conclusiones".
5. COMUNICACION_INTERNA: Memorandos o circulares. Palabras clave: "Memorando", "Circular", "Para:", "De:", "Asunto:", "Cordial saludo".
"""

# Lista auxiliar para validación rápida post-inferencia
CATEGORIAS_VALIDAS = ["QUEJA_RECLAMO", "CONTRATO", "RESOLUCION_ADMINISTRATIVA", 
                      "INFORME_TECNICO", "COMUNICACION_INTERNA"]

# --- FUNCIONES DE UTILIDAD (IMAGEN) ---

def encode_image_to_base64(image: Image.Image) -> str:
    """
    Convierte un objeto PIL Image a una cadena base64 UTF-8.
    
    Args:
        image (Image.Image): Imagen PIL a convertir.
        
    Returns:
        str: Cadena codificada en base64 lista para enviar a la API.
    """
    buffered = io.BytesIO()
    image.save(buffered, format="JPEG")
    return base64.b64encode(buffered.getvalue()).decode('utf-8')

def preprocess_image(image: Image.Image) -> Image.Image:
    """
    Aplica pre-procesamiento a la imagen para mejorar la legibilidad por parte del modelo.
    Realiza dos acciones principales:
    1. Deskewing (Corrección de inclinación): Si el documento está torcido.
    2. Mejora de Contraste: Para resaltar texto tenue.

    Args:
        image (Image.Image): Imagen original extraída del PDF.

    Returns:
        Image.Image: Imagen procesada (corregida y con contraste ajustado).
    """
    try:
        # Conversión para OpenCV
        open_cv_image = np.array(image.convert('RGB'))[:, :, ::-1].copy()
        grayscale = cv2.cvtColor(open_cv_image, cv2.COLOR_BGR2GRAY)
        
        # Detección de ángulo de inclinación
        angle = determine_skew(grayscale)
        
        # Aplicar rotación solo si la inclinación es significativa (> 0.5 grados)
        if angle and abs(angle) > 0.5:
            (h, w) = open_cv_image.shape[:2]
            M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
            rotated = cv2.warpAffine(open_cv_image, M, (w, h), flags=cv2.INTER_CUBIC, borderValue=(255, 255, 255))
            corrected_image = Image.fromarray(cv2.cvtColor(rotated, cv2.COLOR_BGR2RGB))
        else:
            corrected_image = image
        
        # Ajuste de contraste (1.5 es más seguro que 1.8 para evitar saturación en firmas)
        return ImageEnhance.Contrast(corrected_image).enhance(1.5)
    except Exception as e:
        print(f"Advertencia: Falló el preprocesamiento de imagen ({e}). Usando original.")
        return image

# --- LÓGICA CORE DE CLASIFICACIÓN ---

def classify_document_hybrid(image: Image.Image, text_context: str) -> Dict[str, Union[str, float]]:
    """
    Clasifica un documento utilizando un enfoque híbrido (Visión + Texto).
    Utiliza Llama 3.2 Vision a través de Ollama.

    Args:
        image (Image.Image): Imagen de la primera página del documento.
        text_context (str): Texto crudo extraído del PDF (OCR o capa de texto).

    Returns:
        dict: Diccionario con keys {'categoria', 'score', 'evidencia'}.
    """
    base64_image = encode_image_to_base64(image)
    
    # Limpieza de texto: eliminar saltos de línea excesivos para optimizar tokens
    text_snippet = re.sub(r'\s+', ' ', text_context).strip()[:2000].replace('"', "'")

    # Prompt Ingeniería: Inyectamos las definiciones y evitamos sesgo en el score
    prompt_sistema = f"""Eres un clasificador documental experto para una entidad pública.
    
    TU OBJETIVO: Clasificar el documento basándote en la IMAGEN visual y el TEXTO extraído.
    
    DEFINICIONES DE CATEGORÍAS (Criterios estrictos):
    {DEFINICIONES_CATEGORIAS}

    TEXTO EXTRAÍDO DEL DOCUMENTO:
    "{text_snippet}..."

    INSTRUCCIONES DE SALIDA:
    1. Analiza si el documento cumple con las palabras clave visuales o textuales definidas.
    2. Si no encaja claramente en ninguna, usa "REVISION_MANUAL".
    3. 'score' debe representar tu nivel de certeza (0.0 a 1.0).
    
    RESPONDE ÚNICAMENTE CON ESTE FORMATO JSON RAW (Sin Markdown, sin explicaciones previas):
    {{"categoria": "NOMBRE_CATEGORIA", "score": <float_entre_0_y_1>, "evidencia": "breve justificación"}}"""

    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt_sistema},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                        }
                    ],
                }
            ],
            max_tokens=300,
            temperature=0.1, # Baja temperatura para mayor determinismo en JSON
        )
        
        content = response.choices[0].message.content.strip()
        
        # --- Limpieza Robusta de JSON ---
        # 1. Eliminar bloques de código markdown (```json ... ```)
        cleaned = re.sub(r'```(?:json)?|```', '', content).strip()
        # 2. Aislar el objeto JSON eliminando ruido antes y después de las llaves
        cleaned = re.sub(r'^[^{]*', '', cleaned)
        cleaned = re.sub(r'[^}]*$', '', cleaned)
        
        if cleaned:
            result = json.loads(cleaned)
            
            # Normalización y Validación
            cat = result.get("categoria", "").upper().replace(" ", "_")
            
            # Corrección de plurales comunes (alucinaciones leves del modelo)
            if cat.endswith("S") and cat[:-1] in CATEGORIAS_VALIDAS:
                cat = cat[:-1]
                
            if cat not in CATEGORIAS_VALIDAS:
                cat = "REVISION_MANUAL"
            
            # Parsing seguro del Score
            score = result.get("score", 0.5)
            if isinstance(score, str):
                try:
                    score = float(score.replace(",", "."))
                except ValueError:
                    score = 0.5
            score = max(0.0, min(1.0, score)) # Clamping
            
            return {
                "categoria": cat,
                "score": round(score, 2),
                "evidencia": result.get("evidencia", "Sin evidencia detallada")
            }
        else:
            return {"categoria": "REVISION_MANUAL", "score": 0.0, "evidencia": "El modelo no retornó un JSON válido."}
            
    except json.JSONDecodeError as je:
        return {"categoria": "REVISION_MANUAL", "score": 0.0, "evidencia": f"Error de sintaxis JSON: {str(je)[:50]}"}
    except Exception as e:
        return {"categoria": "ERROR", "score": 0.0, "evidencia": f"Fallo en inferencia: {str(e)[:100]}"}

# --- PROCESAMIENTO DE ARCHIVOS ---

def process_pdf(input_path: str) -> dict:
    """
    Versión OPTIMIZADA para documentos nativos digitales.
    Sin deskewing ni filtros de contraste.
    """
    try:
        doc = fitz.open(input_path)
        if len(doc) == 0: 
            return {"categoria": "VACIO", "score": 0.0, "evidencia": "PDF sin páginas"}

        page = doc[0]
        
        # 1. Extracción de Texto (En PDFs digitales esto es 100% preciso, ¡aprovéchalo!)
        text_content = page.get_text("text") or ""
        
        # 2. Renderizado de Imagen (Directo, sin pre-proceso)
        # Nota: DPI=150 sigue siendo bueno. Si la letra es muy pequeña en el digital, 
        # podrías probar DPI=200, pero 150 suele bastar.
        pix = page.get_pixmap(dpi=DPI)
        
        # Convertimos directamente los bytes a una imagen PIL
        img = Image.open(io.BytesIO(pix.tobytes()))
        
        # 3. Clasificación (Pasamos la imagen limpia)
        result = classify_document_hybrid(img, text_content)
        
        doc.close()
        return result
    except Exception as e:
        return {"categoria": "ARCHIVO_CORRUPTO", "score": 0.0, "evidencia": str(e)}

def main():
    """
    Función principal. Itera sobre los archivos en la carpeta de datos y genera el reporte Excel.
    """
    input_folder = "/data" 
    output_excel = "/data/clasificacion_documentos.xlsx"

    pdf_files = glob.glob(os.path.join(input_folder, '*.pdf'))
    
    if not pdf_files:
        print("No se encontraron PDFs en la carpeta 'data'.")
        return

    print(f"Iniciando clasificación de {len(pdf_files)} documentos...")
    
    # Configuración de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificación de Documentos"
    
    # Estilos de Excel
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ["Nombre de Archivo", "Categoría Detectada", "Score", "Justificación"]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Ajuste de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.row_dimensions[1].height = 30
    
    # Loop de procesamiento
    row_num = 2
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        print(f"\n--- Analizando: {filename} ---")
        
        result = process_pdf(pdf_path)
        
        cat = result.get("categoria", "ERROR")
        score = result.get("score", 0.0)
        evidencia = result.get("evidencia", "")
        
        print(f"   -> Resultado: {cat} (Score: {score})")
        
        row_data = [filename, cat, score, evidencia]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        row_num += 1
    
    wb.save(output_excel)
    print(f"\nProceso finalizado. Resultados en: {output_excel}")

if __name__ == "__main__":
    main()