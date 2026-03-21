"""
main_clasificacion_Text.py
==========================

Módulo para la clasificación automática de documentos PDF usando modelos de lenguaje (LLM).

Descripción:
------------
Este script procesa documentos PDF almacenados en una carpeta, extrae su contenido textual
de forma inteligente (priorizando inicio y final del documento), y utiliza un modelo de
lenguaje local (Ollama) para clasificarlos en categorías predefinidas de documentos
administrativos.

Categorías de Clasificación:
----------------------------
    - QUEJA_RECLAMO: Documentos de inconformidad ciudadana
    - CONTRATO: Acuerdos legales entre partes
    - RESOLUCION_ADMINISTRATIVA: Actos administrativos formales
    - INFORME_TECNICO: Reportes de avance, diagnósticos o resultados
    - COMUNICACION_INTERNA: Memorandos o circulares

Flujo de Trabajo:
-----------------
    1. Escanea la carpeta /data buscando archivos PDF
    2. Extrae texto inteligente de cada PDF (primera página, inicio de segunda, última página)
    3. Envía el texto al modelo LLM para clasificación
    4. Genera un reporte Excel con los resultados

Dependencias:
-------------
    - fitz (PyMuPDF): Extracción de texto de PDFs
    - openpyxl: Generación de reportes Excel
    - openai: Cliente para comunicación con Ollama (API compatible)

Configuración Requerida:
------------------------
    - Ollama ejecutándose con el modelo configurado (por defecto: qwen2.5:7b)
    - Variable de entorno OLLAMA_HOST (opcional, default: http://host.docker.internal:11434/v1)

Uso:
----
    $ python main_clasificacion_Text.py

    Los PDFs deben estar en la carpeta /data y el reporte se generará como
    /data/clasificacion_documentos_texto.xlsx

Autor: Javier Mauricio Sierra
Fecha: Enero 2026
Versión: 1.0.0
"""

import os
import fitz  # PyMuPDF - Biblioteca para manipulación de documentos PDF
import re
import json
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openai import OpenAI  # Cliente OpenAI compatible con Ollama
from typing import Dict, Any, Union

# =============================================================================
# CONFIGURACIÓN DEL MODELO LLM
# =============================================================================
# Configuración para conectar con Ollama (servidor LLM local).
# Modelos recomendados: qwen2.5:7b, qwen2.5-coder:7b
# Para instalar un modelo: `ollama pull qwen2.5:7b`

#MODEL_NAME = "qwen2.5:7b"  
MODEL_NAME = "qwen2.5-coder:7b"  # Modelo de lenguaje a utilizar para clasificación
OLLAMA_URL = os.getenv('OLLAMA_HOST', 'http://host.docker.internal:11434/v1')  # URL del servidor Ollama

# Cliente OpenAI configurado para comunicarse con Ollama
# Nota: Ollama expone una API compatible con OpenAI
client = OpenAI(base_url=OLLAMA_URL, api_key='ollama')

# =============================================================================
# DEFINICIONES DE NEGOCIO - CATEGORÍAS DE CLASIFICACIÓN
# =============================================================================
# Estas definiciones se envían como contexto al modelo LLM para que pueda
# identificar el tipo de documento basándose en palabras clave específicas.
# 
# IMPORTANTE: Modificar estas definiciones afecta directamente la precisión
# de la clasificación. Agregar más palabras clave mejora la detección.

DEFINICIONES_CATEGORIAS = """
1. QUEJA_RECLAMO: Documentos donde un ciudadano expresa inconformidad. Palabras clave: "Petición", "Reclamo", "Solicito", "Inconformidad", "Derecho de petición".
2. CONTRATO: Acuerdos legales entre partes. Palabras clave: "Contratante", "Contratista", "Cláusulas", "Objeto del contrato", "Prestación de servicios", "Firma".
3. RESOLUCION_ADMINISTRATIVA: Actos administrativos formales. Palabras clave: "RESOLUCION No", "RESUELVE", "CONSIDERANDO", "ARTICULO PRIMERO", "COMUNIQUESE Y CUMPLASE".
4. INFORME_TECNICO: Reportes de avance, diagnósticos o resultados. Palabras clave: "Informe de gestión", "Resultados", "Diagnóstico", "Objetivo", "Alcance", "Conclusiones".
5. COMUNICACION_INTERNA: Memorandos o circulares. Palabras clave: "Memorando", "Circular", "Para:", "De:", "Asunto:", "Cordial saludo".
"""

# Lista de categorías válidas para validación de respuestas del LLM
# Cualquier categoría no reconocida se marca como REVISION_MANUAL
CATEGORIAS_VALIDAS = [
    "QUEJA_RECLAMO", 
    "CONTRATO", 
    "RESOLUCION_ADMINISTRATIVA", 
    "INFORME_TECNICO", 
    "COMUNICACION_INTERNA"
]

# =============================================================================
# FUNCIONES DE EXTRACCIÓN DE TEXTO
# =============================================================================

def extract_text_smart(doc: fitz.Document, max_chars: int = 4000) -> str:
    """
    Extrae texto de un documento PDF de forma inteligente para clasificación.
    
    Implementa una estrategia optimizada que prioriza las secciones del documento
    donde típicamente se encuentra la información más relevante para determinar
    el tipo de documento:
    
    - **Primera página**: Contiene título, encabezados y contexto inicial
    - **Inicio de segunda página**: Proporciona contexto adicional
    - **Última página**: Incluye firmas, sellos y conclusiones
    
    Esta estrategia reduce significativamente el uso de tokens del LLM mientras
    mantiene alta precisión en la clasificación.
    
    Args:
        doc (fitz.Document): Documento PDF abierto con PyMuPDF.
        max_chars (int, optional): Máximo de caracteres a extraer. Default: 4000.
            Este límite previene el envío de textos excesivamente largos al LLM.
    
    Returns:
        str: Texto extraído y limpiado, listo para enviar al clasificador.
    
    Example:
        >>> doc = fitz.open("documento.pdf")
        >>> texto = extract_text_smart(doc, max_chars=3000)
        >>> print(len(texto))  # <= 3000
    
    Note:
        Si el documento es una imagen escaneada sin OCR, retornará un texto
        muy corto o vacío. Esto se detecta en process_pdf() para marcarlo
        como REVISION_MANUAL.
    """
    full_text = ""
    total_pages = len(doc)
    
    # PASO 1: Extraer primera página completa
    # Contiene: título, encabezado institucional, número de resolución, etc.
    if total_pages > 0:
        full_text += doc[0].get_text("text") + "\n"
        
    # PASO 2: Extraer inicio de segunda página (máximo 500 caracteres)
    # Proporciona contexto adicional sin consumir demasiados tokens
    if total_pages > 1:
        full_text += doc[1].get_text("text")[:500] + "\n"
        
    # PASO 3: Extraer última página para documentos largos (>2 páginas)
    # Captura: firmas, sellos, conclusiones, pie de página institucional
    if total_pages > 2:
        full_text += "\n...[CONTENIDO OMITIDO]...\n"  # Indicador para el LLM
        full_text += doc[total_pages - 1].get_text("text")

    # LIMPIEZA: Normalizar espacios en blanco para optimizar tokens
    # Convierte múltiples espacios/saltos de línea en un solo espacio
    cleaned_text = re.sub(r'\s+', ' ', full_text).strip()
    
    return cleaned_text[:max_chars]

# =============================================================================
# LÓGICA CORE DE CLASIFICACIÓN (TEXT ONLY)
# =============================================================================

def classify_document_text_only(text_content: str) -> Dict[str, Union[str, float]]:
    """
    Clasifica un documento basándose únicamente en su contenido textual.
    
    Envía el texto extraído al modelo LLM junto con las definiciones de categorías,
    y procesa la respuesta JSON para obtener la clasificación.
    
    Args:
        text_content (str): Texto extraído del documento PDF a clasificar.
            Debe estar limpio y con longitud razonable (< 4000 caracteres).
    
    Returns:
        Dict[str, Union[str, float]]: Diccionario con el resultado de clasificación:
            - "categoria" (str): Categoría asignada (ej: "CONTRATO", "REVISION_MANUAL")
            - "score" (float): Nivel de confianza entre 0.0 y 1.0
            - "evidencia" (str): Palabras clave encontradas que justifican la clasificación
    
    Example:
        >>> texto = "RESOLUCION No. 123 Por medio de la cual RESUELVE..."
        >>> result = classify_document_text_only(texto)
        >>> print(result)
        {"categoria": "RESOLUCION_ADMINISTRATIVA", "score": 0.95, "evidencia": "RESOLUCION No, RESUELVE"}
    
    Posibles categorías de retorno:
        - Categorías válidas: Las definidas en CATEGORIAS_VALIDAS
        - REVISION_MANUAL: Cuando no se puede determinar la categoría o el JSON es inválido
        - ERROR: Cuando ocurre una excepción durante el procesamiento
    
    Note:
        - Usa temperature=0.0 para resultados determinísticos
        - Usa seed=42 para reproducibilidad entre ejecuciones
    """
    
    # Construcción del prompt con instrucciones detalladas para el LLM
    prompt_sistema = f"""Eres un experto analista documental.
    
    OBJETIVO: Clasificar el texto proporcionado según las siguientes definiciones estrictas.
    
    DEFINICIONES:
    {DEFINICIONES_CATEGORIAS}

    CONTENIDO DEL DOCUMENTO:
    "{text_content}"

    INSTRUCCIONES:
    1. Busca palabras clave específicas de las definiciones.
    2. Determina la categoría más probable.
    3. Asigna un score de confianza (0.0 a 1.0).
    
    RESPONDE ÚNICAMENTE CON ESTE FORMATO JSON RAW (Sin Markdown):
    {{"categoria": "NOMBRE_CATEGORIA", "score": <float>, "evidencia": "palabras clave encontradas"}}"""

    try:
        # Llamada al modelo LLM vía API compatible con OpenAI
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {
                    "role": "system", 
                    "content": "Eres un asistente administrativo experto que SOLO habla en JSON. No incluyas explicaciones, markdown ni texto adicional."
                },
                {"role": "user", "content": prompt_sistema}
            ],
            max_tokens=300,       # Suficiente para el JSON de respuesta
            temperature=0.0,      # Determinismo máximo para consistencia
            seed=42,              # Semilla fija para reproducibilidad
        )
        
        content = response.choices[0].message.content.strip()
        
        # ==== LIMPIEZA Y PARSING DEL JSON ====
        # El LLM a veces envuelve el JSON en bloques de código markdown
        cleaned = re.sub(r'```(?:json)?|```', '', content).strip()
        # Eliminar cualquier texto antes del primer '{'
        cleaned = re.sub(r'^[^{]*', '', cleaned)
        # Eliminar cualquier texto después del último '}'
        cleaned = re.sub(r'[^}]*$', '', cleaned)
        
        if cleaned:
            result = json.loads(cleaned)
            
            # ==== NORMALIZACIÓN DE CATEGORÍA ====
            # Convertir a mayúsculas y reemplazar espacios por guiones bajos
            cat = result.get("categoria", "").upper().replace(" ", "_")
            # Manejar plurales incorrectos (ej: "CONTRATOS" -> "CONTRATO")
            if cat.endswith("S") and cat[:-1] in CATEGORIAS_VALIDAS: 
                cat = cat[:-1]
            # Validar que sea una categoría conocida
            if cat not in CATEGORIAS_VALIDAS: 
                cat = "REVISION_MANUAL"
            
            # ==== NORMALIZACIÓN DEL SCORE ====
            score = result.get("score", 0.5)
            if isinstance(score, str):
                # Manejar scores como string (ej: "0,85" con coma decimal)
                try: 
                    score = float(score.replace(",", "."))
                except: 
                    score = 0.5
            
            return {
                "categoria": cat,
                "score": max(0.0, min(1.0, score)),  # Asegurar rango [0.0, 1.0]
                "evidencia": result.get("evidencia", "Sin evidencia")
            }
        else:
            # JSON vacío o malformado
            return {"categoria": "REVISION_MANUAL", "score": 0.0, "evidencia": "JSON ilegible"}
            
    except Exception as e:
        # Capturar cualquier error (conexión, timeout, parsing, etc.)
        return {"categoria": "ERROR", "score": 0.0, "evidencia": str(e)[:100]}

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DE PDF
# =============================================================================

def process_pdf(input_path: str) -> dict:
    """
    Procesa un archivo PDF individual y retorna su clasificación.
    
    Esta función actúa como orquestador del pipeline de clasificación:
    1. Abre el archivo PDF
    2. Valida que tenga contenido
    3. Extrae el texto de forma inteligente
    4. Envía el texto al clasificador LLM
    5. Retorna el resultado
    
    Args:
        input_path (str): Ruta absoluta al archivo PDF a procesar.
    
    Returns:
        dict: Diccionario con el resultado de clasificación:
            - "categoria" (str): Categoría asignada al documento
            - "score" (float): Nivel de confianza (0.0 - 1.0)
            - "evidencia" (str): Justificación de la clasificación
    
    Categorías especiales de error:
        - VACIO: El PDF no tiene páginas
        - REVISION_MANUAL: PDF es imagen escaneada sin OCR (< 50 caracteres)
        - ARCHIVO_CORRUPTO: Error al abrir o procesar el archivo
    
    Example:
        >>> result = process_pdf("/data/contrato_2024.pdf")
        >>> print(result)
        {"categoria": "CONTRATO", "score": 0.92, "evidencia": "Contratante, Cláusulas"}
    """
    try:
        # Abrir el documento PDF con PyMuPDF
        doc = fitz.open(input_path)
        
        # Validación: Documento sin páginas
        if len(doc) == 0: 
            return {"categoria": "VACIO", "score": 0.0, "evidencia": "Sin páginas"}

        # Extracción inteligente de texto (Inicio + Final del documento)
        text_content = extract_text_smart(doc)
        
        # Validación: PDF es imagen escaneada (sin capa de texto extraíble)
        # Si el texto extraído es muy corto, probablemente es un escaneo
        if len(text_content) < 50:
             doc.close()
             return {
                 "categoria": "REVISION_MANUAL", 
                 "score": 0.0, 
                 "evidencia": "PDF es imagen (sin capa de texto)"
             }

        # Clasificación mediante LLM
        result = classify_document_text_only(text_content)
        
        # Cerrar el documento para liberar recursos
        doc.close()
        return result
        
    except Exception as e:
        # Capturar errores de archivo (corrupto, protegido, formato inválido)
        return {"categoria": "ARCHIVO_CORRUPTO", "score": 0.0, "evidencia": str(e)}

# =============================================================================
# FUNCIÓN PRINCIPAL - PUNTO DE ENTRADA
# =============================================================================

def main():
    """
    Función principal que ejecuta el pipeline completo de clasificación.
    
    Flujo de ejecución:
        1. Escanea la carpeta /data buscando archivos *.pdf
        2. Procesa cada PDF secuencialmente
        3. Muestra progreso en consola
        4. Genera reporte Excel con resultados
    
    Configuración:
        - Carpeta de entrada: /data (configurable en input_folder)
        - Archivo de salida: /data/clasificacion_documentos_texto.xlsx
    
    Estructura del reporte Excel:
        - Columna A: Nombre de Archivo
        - Columna B: Categoría asignada
        - Columna C: Score de confianza (0.0 - 1.0)
        - Columna D: Evidencia/Palabras clave encontradas
    
    Example:
        $ python main_clasificacion_Text.py
        Iniciando clasificación TEXT-ONLY de 15 documentos...
        Analizando: contrato_001.pdf... [CONTRATO]
        Analizando: resolucion_234.pdf... [RESOLUCION_ADMINISTRATIVA]
        ...
        Finalizado. Reporte en: /data/clasificacion_documentos_texto.xlsx
    """
    # Configuración de rutas
    input_folder = "/data"  # Carpeta donde se encuentran los PDFs
    output_excel = "/data/clasificacion_documentos_texto.xlsx"  # Ruta del reporte

    # Buscar todos los archivos PDF en la carpeta
    pdf_files = glob.glob(os.path.join(input_folder, '*.pdf'))
    
    # Validación: No hay PDFs para procesar
    if not pdf_files:
        print("No se encontraron PDFs.")
        return

    print(f"Iniciando clasificación TEXT-ONLY de {len(pdf_files)} documentos...")
    
    # ========================================================================
    # CONFIGURACIÓN DEL ARCHIVO EXCEL
    # ========================================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificación Texto"
    
    # Definir encabezados de columnas
    headers = ["Nombre de Archivo", "Categoría", "Score", "Evidencia"]
    
    # Estilos para el encabezado
    header_font = Font(bold=True, color='FFFFFF')  # Texto blanco en negrita
    header_fill = PatternFill(start_color='4472C4', fill_type='solid')  # Fondo azul
    
    # Aplicar encabezados con estilo
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    # Configurar anchos de columna para mejor legibilidad
    ws.column_dimensions['A'].width = 40   # Nombre de archivo
    ws.column_dimensions['B'].width = 30   # Categoría
    ws.column_dimensions['D'].width = 60   # Evidencia
    ws.row_dimensions[1].height = 30       # Altura del encabezado
    
    # ========================================================================
    # PROCESAMIENTO DE DOCUMENTOS
    # ========================================================================
    row_num = 2  # Comenzar en la fila 2 (después del encabezado)
    
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        print(f"Analizando: {filename}...", end=" ")
        
        # Procesar el PDF y obtener clasificación
        result = process_pdf(pdf_path)
        
        # Mostrar resultado en consola
        print(f"[{result.get('categoria')}]")
        
        # Escribir resultado en Excel
        row_data = [
            filename, 
            result.get("categoria"), 
            result.get("score"), 
            result.get("evidencia")
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)
        
        # Establecer altura de la fila de datos
        ws.row_dimensions[row_num].height = 50
        
        row_num += 1
    
    # ========================================================================
    # GUARDAR REPORTE
    # ========================================================================
    wb.save(output_excel)
    print(f"\nFinalizado. Reporte en: {output_excel}")


# =============================================================================
# PUNTO DE ENTRADA DEL SCRIPT
# =============================================================================
if __name__ == "__main__":
    main()