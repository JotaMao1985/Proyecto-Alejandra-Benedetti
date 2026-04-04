"""
core.py — Funciones reutilizables del pipeline SAM 3 + VLM.

Extraídas de analisis_estanteria_hybrid.py para ser usadas tanto
por el CLI como por la webapp Flask.
"""

import io
import os
import json
import base64
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import numpy as np
from PIL import Image
import torch
from openai import OpenAI

# =============================================================================
# Configuración
# =============================================================================
SAM3_CONFIDENCE = 0.1
SAM3_TEXT_PROMPT = "product"

OLLAMA_URL = os.environ.get("OLLAMA_HOST", "http://host.docker.internal:11434/v1")
VLM_MODEL = "qwen2.5vl:7b"


# =============================================================================
# Data Classes
# =============================================================================
@dataclass
class ProductResult:
    name: str
    price: Optional[float]
    quantity: int = 1
    confidence: float = 0.0
    mask_id: int = 0
    bbox: tuple[int, int, int, int] = (0, 0, 0, 0)


@dataclass
class ShelfAnalysisResult:
    image_name: str
    products: list[ProductResult] = field(default_factory=list)
    total_products: int = 0
    total_unique_products: int = 0
    sam_masks_generated: int = 0
    vlm_calls: int = 0
    errors: list[str] = field(default_factory=list)
    processing_time: float = 0.0


# =============================================================================
# SAM 3
# =============================================================================
def get_sam_device() -> str:
    """Detecta el device disponible: cuda > mps > cpu."""
    if torch.cuda.is_available():
        return "cuda"
    if torch.backends.mps.is_available():
        return "mps"
    return "cpu"


def load_sam3_model(device=None, bpe_path=None, confidence_threshold=SAM3_CONFIDENCE):
    """Carga SAM 3 y retorna el Sam3Processor."""
    from sam3 import build_sam3_image_model
    from sam3.model.sam3_image_processor import Sam3Processor

    if device is None:
        device = get_sam_device()

    if device == "cuda" and not torch.cuda.is_available():
        print("[WARN] CUDA no disponible, usando CPU")
        device = "cpu"

    print(f"[INFO] SAM 3 usando device: {device}")

    build_kwargs = {"device": device}
    if bpe_path:
        build_kwargs["bpe_path"] = bpe_path

    model = build_sam3_image_model(**build_kwargs)

    if device in ("mps", "cuda"):
        model = model.to(device)

    processor = Sam3Processor(model, confidence_threshold=confidence_threshold)
    return processor


def generate_sam3_masks(processor, image, text_prompt=SAM3_TEXT_PROMPT,
                        min_area=500, max_masks=100):
    """Genera máscaras de productos usando SAM 3 con text prompt."""
    if torch.cuda.is_available():
        autocast_ctx = torch.autocast("cuda", dtype=torch.bfloat16)
    elif torch.backends.mps.is_available():
        autocast_ctx = torch.autocast("mps", dtype=torch.float16)
    else:
        autocast_ctx = torch.autocast("cpu", dtype=torch.bfloat16)

    with autocast_ctx:
        state = processor.set_image(image)
        state = processor.set_text_prompt(text_prompt, state)

    masks = state.get("masks", [])
    boxes = state.get("boxes", [])
    scores = state.get("scores", [])

    if masks is None or len(masks) == 0:
        return []

    filtered = []
    for i, (mask_tensor, box, score) in enumerate(zip(masks, boxes, scores)):
        mask_np = mask_tensor[0].cpu().numpy().astype(bool)
        area = mask_np.sum()
        if area < min_area:
            continue

        score_val = float(score)
        x1, y1, x2, y2 = [int(v) for v in box[:4]]
        filtered.append((mask_np, (x1, y1, x2, y2), score_val))

        if len(filtered) >= max_masks:
            break

    return filtered


def crop_mask_region(image, mask, bbox, margin=10):
    """Recorta la región de la imagen correspondiente a un mask."""
    x1, y1, x2, y2 = bbox
    h, w = mask.shape

    x1_c = max(0, x1 - margin)
    y1_c = max(0, y1 - margin)
    x2_c = min(w, x2 + margin)
    y2_c = min(h, y2 + margin)

    mask_crop = mask[y1_c:y2_c, x1_c:x2_c]
    if not mask_crop.any():
        return None

    pil_crop = image.crop((x1_c, y1_c, x2_c, y2_c))
    mask_img = Image.fromarray((mask_crop * 255).astype(np.uint8), mode="L")
    mask_resized = mask_img.resize(pil_crop.size, Image.NEAREST)

    result = Image.new("RGBA", pil_crop.size, (0, 0, 0, 0))
    result.paste(pil_crop, (0, 0))
    result.putalpha(mask_resized)
    result_rgb = result.convert("RGB")

    alpha_arr = np.array(mask_resized)
    if alpha_arr.max() == 0:
        return None

    mask_bool_u = alpha_arr > 127
    rows = np.any(mask_bool_u, axis=1)
    cols = np.any(mask_bool_u, axis=0)
    if not rows.any() or not cols.any():
        return None

    rmin, rmax = np.where(rows)[0][[0, -1]]
    cmin, cmax = np.where(cols)[0][[0, -1]]
    return result_rgb.crop((cmin, rmin, cmax + 1, rmax + 1))


# =============================================================================
# VLM
# =============================================================================
SYSTEM_PROMPT_VLM = """Eres un asistente que identifica productos en imágenes de estanterías de supermercado.

Devuelve SOLO un JSON con esta estructura exacta, sin texto adicional:
{"name": "nombre del producto", "price": 12.99, "confidence": 0.95}

Reglas:
- name: nombre corto del producto (máx 5 palabras)
- price: precio del producto en la etiqueta visible, o null si no se ve
- confidence: número entre 0.0 y 1.0 indicando qué tan seguro estás
- Si no hay producto válido en la imagen, devuelve: {"name": "desconocido", "price": null, "confidence": 0.1}
- NO cuentes productos, solo identifica ESTE producto individual
- Sé conciso, el nombre debe tener máximo 5 palabras"""


def encode_image_to_base64(image):
    """Convierte PIL Image a base64 JPEG."""
    buffer = io.BytesIO()
    image.save(buffer, format="JPEG", quality=85)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def classify_single_product(image, client, model=VLM_MODEL, temperature=0.1):
    """Clasifica un solo producto (recortado por SAM) usando el VLM."""
    img_b64 = encode_image_to_base64(image)
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT_VLM},
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                        {"type": "text", "text": "Identifica este producto de la estantería."},
                    ],
                },
            ],
            temperature=temperature,
            max_tokens=150,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```json"):
            raw = raw[7:]
        if raw.startswith("```"):
            raw = raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3]

        data = json.loads(raw.strip())
        return ProductResult(
            name=data.get("name", "desconocido"),
            price=data.get("price"),
            quantity=1,
            confidence=data.get("confidence", 0.5),
        )
    except Exception as e:
        print(f"  [WARN] Error clasificando producto: {type(e).__name__}: {e}")
        return ProductResult(name="error_parse", price=None, quantity=1, confidence=0.0)


# =============================================================================
# Pipeline
# =============================================================================
def analyze_shelf(image_path, sam_processor, vlm_client, text_prompt=SAM3_TEXT_PROMPT,
                  max_workers=1, min_mask_area=500, vlm_model=VLM_MODEL,
                  progress_callback=None):
    """
    Pipeline híbrido SAM 3 + VLM para una imagen.

    Args:
        progress_callback: callable(procesadas, total) para reportar progreso.
    """
    start_time = time.time()
    image_name = Path(image_path).name
    result = ShelfAnalysisResult(image_name=image_name)
    errors = []

    try:
        image = Image.open(image_path).convert("RGB")
    except Exception as e:
        errors.append(f"Error al abrir imagen: {e}")
        result.errors = errors
        return result

    masks_with_data = generate_sam3_masks(
        sam_processor, image, text_prompt=text_prompt,
        min_area=min_mask_area, max_masks=100,
    )
    result.sam_masks_generated = len(masks_with_data)

    if not masks_with_data:
        errors.append("SAM 3 no detectó ningún producto")
        result.errors = errors
        result.processing_time = time.time() - start_time
        return result

    def process_single_mask(args):
        idx, (mask, bbox, score) = args
        crop = crop_mask_region(image, mask, bbox, margin=10)
        if crop is None:
            return None
        vlm_result = classify_single_product(crop, vlm_client, model=vlm_model)
        vlm_result.mask_id = idx
        vlm_result.bbox = bbox
        vlm_result.confidence = max(vlm_result.confidence, score)
        return vlm_result

    vlm_results = []
    vlm_calls = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(process_single_mask, (i, md)): i
            for i, md in enumerate(masks_with_data)
        }
        for future in as_completed(futures):
            vlm_calls += 1
            try:
                res = future.result()
                if res is not None:
                    vlm_results.append(res)
            except Exception as e:
                errors.append(f"Error en mask {futures[future]}: {e}")
            if progress_callback:
                progress_callback(vlm_calls, len(masks_with_data))

    result.products = vlm_results
    result.total_products = len(vlm_results)
    result.vlm_calls = vlm_calls
    result.errors = errors
    result.processing_time = time.time() - start_time
    return result


def deduplicate_products(products):
    """Deduplica productos por nombre + precio."""
    if not products:
        return []

    deduped = []
    used = set()
    products_sorted = sorted(products, key=lambda p: p.confidence, reverse=True)

    for p in products_sorted:
        if p.name in ("desconocido", "error_parse"):
            continue

        is_duplicate = False
        p_name_norm = p.name.lower().strip()

        for used_idx in used:
            other = deduped[used_idx]
            other_norm = other.name.lower().strip()
            name_similar = p_name_norm in other_norm or other_norm in p_name_norm
            price_match = (
                p.price is not None and other.price is not None
                and abs(p.price - other.price) < 0.01
            )
            if name_similar and price_match:
                is_duplicate = True
                break

        if not is_duplicate:
            deduped.append(p)
            used.add(len(deduped) - 1)

    return deduped


def export_to_excel(results, output_path):
    """Exporta resultados a Excel con 3 hojas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_detalle = wb.create_sheet("Detalle de Productos")
    ws_stats = wb.create_sheet("Estadísticas")

    hfill = PatternFill(start_color="3D008D", end_color="ED1E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(*(Side(style="thin") for _ in range(4)))

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill, c.font, c.alignment, c.border = hfill, hfont, halign, border

    # Resumen
    write_header(ws_summary, ["Imagen", "Productos", "Masks SAM", "Llamadas VLM", "Tiempo (s)", "Errores"])
    for col, w in zip("ABCDEF", [20, 12, 12, 15, 15, 12]):
        ws_summary.column_dimensions[col].width = w
    for row, r in enumerate(results, 2):
        vals = [r.image_name, r.total_unique_products, r.sam_masks_generated,
                r.vlm_calls, round(r.processing_time, 1),
                ", ".join(r.errors) if r.errors else "-"]
        for col, v in enumerate(vals, 1):
            ws_summary.cell(row=row, column=col, value=v).border = border

    # Detalle
    write_header(ws_detalle, ["Imagen", "Producto", "Precio", "Cantidad", "Confianza", "Mask ID"])
    for col in range(1, 7):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 18
    row = 2
    for r in results:
        for p in r.products:
            vals = [r.image_name, p.name, p.price, p.quantity, round(p.confidence, 2), p.mask_id]
            for col, v in enumerate(vals, 1):
                ws_detalle.cell(row=row, column=col, value=v).border = border
            row += 1

    # Estadísticas
    write_header(ws_stats, ["Métrica", "Valor"])
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 15
    total_imgs = len(results)
    total_products = sum(r.total_unique_products for r in results)
    total_masks = sum(r.sam_masks_generated for r in results)
    total_vlm = sum(r.vlm_calls for r in results)
    avg_time = sum(r.processing_time for r in results) / total_imgs if total_imgs else 0
    stats = [
        ("Total imágenes", total_imgs), ("Total productos", total_products),
        ("Total masks SAM", total_masks), ("Total llamadas VLM", total_vlm),
        ("Tiempo promedio (s)", round(avg_time, 1)),
        ("Productos promedio/imagen", round(total_products / total_imgs, 1) if total_imgs else 0),
    ]
    for i, (m, v) in enumerate(stats, 2):
        ws_stats.cell(row=i, column=1, value=m).border = border
        ws_stats.cell(row=i, column=2, value=v).border = border

    wb.save(output_path)
